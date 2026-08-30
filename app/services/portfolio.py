from __future__ import annotations

import csv
import io
import json
import re
import threading
import uuid
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
_LOCK = threading.Lock()

def _get_user_dir(username: str | None = None) -> Path:
    from app.services.user_manager import get_user_data_dir
    return get_user_data_dir(username)

def _get_portfolio_file(username: str | None = None) -> Path:
    return _get_user_dir(username) / "portfolio.json"

EMPTY_PORTFOLIO: dict[str, Any] = {
    "settings": {"fx_rates": {"KRW": 1.0}, "fx_info": {}, "daily_snapshot": {}, "cash_balances": {}},
    "accounts": [],
    "holdings": [],
    "bank_accounts": [],
    "savings_accounts": [],
    "insurance_accounts": [],
    "loan_accounts": [],
    "real_estates": [],
    "updated_at": None,
}

FIELD_ALIASES = {
    "broker": ("증권사", "broker", "brokerage"),
    "account": ("계좌명", "계좌번호", "계좌", "account", "account_name"),
    "owner": ("소유자", "소유자명", "가족", "owner", "member"),
    "code": ("종목코드", "종목번호", "단축코드", "code", "symbol", "ticker"),
    "name": ("종목명", "종목", "name", "security_name"),
    "quantity": ("보유수량", "수량", "잔고수량", "quantity", "shares"),
    "avg_price": ("평균매입가", "매입단가", "매입평균가", "buy_price", "avg_price"),
    "current_price": ("현재가", "평가단가", "price", "current_price", "last_price"),
    "currency": ("통화", "currency", "화폐"),
    "market": ("거래소", "시장", "market", "exchange"),
}


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _ensure_data_file(username: str | None = None) -> Path:
    f = _get_portfolio_file(username)
    f.parent.mkdir(parents=True, exist_ok=True)
    if not f.exists():
        f.write_text(json.dumps(EMPTY_PORTFOLIO, ensure_ascii=False, indent=2), encoding="utf-8")
    return f


def read_portfolio(username: str | None = None) -> dict[str, Any]:
    with _LOCK:
        f = _ensure_data_file(username)
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            data = deepcopy(EMPTY_PORTFOLIO)
        data.setdefault("settings", deepcopy(EMPTY_PORTFOLIO["settings"]))
        data["settings"].setdefault("fx_rates", {"KRW": 1.0})
        data["settings"]["fx_rates"].setdefault("KRW", 1.0)
        data["settings"].setdefault("fx_info", {})
        data["settings"].setdefault("daily_snapshot", {})
        data["settings"].setdefault("cash_balances", {})
        data.setdefault("accounts", [])
        data.setdefault("holdings", [])
        data.setdefault("updated_at", None)
        previous_date = str(data.get("updated_at") or "")[:10]
        today = datetime.now().astimezone().date().isoformat()
        if previous_date and previous_date != today and data["settings"].get("daily_snapshot", {}).get("date") != previous_date:
            previous_value = 0.0
            for holding in data["holdings"]:
                currency = normalize_currency(holding.get("currency"))
                rate = to_number(data["settings"]["fx_rates"].get(currency), 1.0 if currency == "KRW" else 0.0)
                previous_value += to_number(holding.get("quantity")) * to_number(holding.get("current_price")) * rate
            data["settings"]["daily_snapshot"] = {"date": previous_date, "value_krw": previous_value}
        return data


def write_portfolio(data: dict[str, Any], username: str | None = None) -> dict[str, Any]:
    with _LOCK:
        f = _ensure_data_file(username)
        data["updated_at"] = now_iso()
        temp_file = f.with_suffix(".json.tmp")
        temp_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_file.replace(f)
        return data


def migrate_add_family_group(username: str | None = None) -> None:
    """Ensure all account entries have a 'family_group' key.
    Existing accounts without the key will get the default value 'All'."""
    data = read_portfolio(username)
    modified = False
    for account in data.get('accounts', []):
        if 'family_group' not in account:
            account['family_group'] = 'All'
            modified = True
    if modified:
        write_portfolio(data, username)



def to_number(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value).replace(",", ""))
    try:
        return float(cleaned) if cleaned not in {"", "-", "."} else default
    except ValueError:
        return default


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_currency(value: Any) -> str:
    value = clean_text(value).upper()
    return {"원": "KRW", "₩": "KRW", "달러": "USD", "미국달러": "USD", "$": "USD", "US$": "USD"}.get(value, value or "KRW")


def normalize_market(value: Any, code: str, currency: str) -> str:
    value = clean_text(value).upper().replace(" ", "")
    aliases = {"코스피": "KRX", "코스닥": "KRX", "국내": "KRX", "NASDAQ": "NAS", "나스닥": "NAS", "NYSE": "NYS", "뉴욕": "NYS", "AMEX": "AMX", "아멕스": "AMX"}
    if value in aliases:
        return aliases[value]
    if value:
        return value
    return "KRX" if currency == "KRW" and code.isdigit() else ""


def get_or_add_account(data: dict[str, Any], broker: str, account_name: str, family_group: str = "All", source: str = "import") -> str:
    broker = broker or "기타 증권사"
    account_name = account_name or f"{broker} 계좌"
    for account in data["accounts"]:
        if account["broker"] == broker and account["name"] == account_name:
            return account["id"]
    account_id = str(uuid.uuid4())
    data["accounts"].append({
        "id": account_id,
        "broker": broker,
        "name": account_name,
        "family_group": family_group,
        "source": source,
    })
    return account_id



DEFAULT_SECTOR_MAP = {
    # 반도체
    "005930": "반도체", "005935": "반도체", "000660": "반도체", "0193T0": "반도체", "0193W0": "반도체",
    "091160": "반도체", "240810": "반도체", "353200": "반도체", "395160": "반도체", "471990": "반도체",
    "NVDA": "반도체", "AVGO": "반도체", "TSM": "반도체", "NVDL": "반도체",

    # IT·빅테크
    "AAPL": "IT·빅테크", "MSFT": "IT·빅테크", "035420": "IT·빅테크", "030000": "IT·빅테크",

    # 2차전지·모빌리티
    "364980": "2차전지", "TSLA": "모빌리티·2차전지", "005380": "자동차·운송",

    # 금융·지주
    "0089D0": "금융·지주", "039490": "금융·지주", "055550": "금융·지주", "086790": "금융·지주",
    "091170": "금융·지주", "102970": "금융·지주", "105560": "금융·지주", "138040": "금융·지주", "316140": "금융·지주",

    # 전력·인프라·건설
    "015760": "전력·인프라", "117700": "건설·인프라", "267260": "전력·인프라",
    "487130": "전력·인프라", "487240": "전력·인프라",

    # 방산·조선
    "449450": "방산·우주", "466920": "조선·기계",

    # 바이오·헬스케어
    "244580": "바이오·헬스케어",

    # 소비재·엔터·유통
    "228790": "소비재·뷰티", "475050": "엔터·미디어", "KO": "음료·소비재", "WMT": "도소매·유통",

    # 리츠·부동산
    "481850": "리츠·부동산",

    # 미국 대표지수·ETF
    "QQQM": "미국 대표지수", "SPYG": "미국 대표지수", "QLD": "미국 대표지수", "TQQQ": "미국 대표지수",
    "QNDX": "미국 대표지수", "IVV": "미국 대표지수", "0015B0": "미국 대표지수", "0026S0": "미국 대표지수",
    "0069M0": "미국 대표지수", "0104H0": "미국 대표지수", "0190M0": "미국 대표지수", "379810": "미국 대표지수",

    # 국내 대표지수·ETF
    "069500": "국내 대표지수", "0163Y0": "국내 대표지수", "0088N0": "국내 대표지수",

    # 채권·안전자산
    "TLT": "채권·안전자산",
}


def get_default_sector(code: str, name: str = "") -> str:
    c = str(code).strip().upper()
    if c in DEFAULT_SECTOR_MAP:
        return DEFAULT_SECTOR_MAP[c]
    n = name.upper()
    if any(k in n for k in ["반도체", "SEMICONDUCTOR", "CHIP"]): return "반도체"
    if any(k in n for k in ["2차전지", "배터리", "BATTERY"]): return "2차전지"
    if any(k in n for k in ["금융", "은행", "증권", "지주", "FINANCIAL", "BANK"]): return "금융·지주"
    if any(k in n for k in ["전력", "인프라", "에너지", "ENERGY", "POWER"]): return "전력·인프라"
    if any(k in n for k in ["바이오", "헬스케어", "PHARMA", "BIO", "HEALTH"]): return "바이오·헬스케어"
    if any(k in n for k in ["나스닥", "S&P", "다우", "INDEX", "200", "코스닥"]): return "대표지수·ETF"
    return "기타"


def normalize_holding(raw: dict[str, Any], account_id: str, broker: str, account_name: str, source: str = "manual") -> dict[str, Any]:
    from app.services.stock_master import resolve_stock_info
    raw_code = clean_text(raw.get("code"))
    raw_name = clean_text(raw.get("name"))
    raw_curr = normalize_currency(raw.get("currency"))

    code, name, currency = resolve_stock_info(raw_code, raw_name, raw_curr)
    if not name:
        name = code
    return {
        "id": clean_text(raw.get("id")) or str(uuid.uuid4()),
        "account_id": account_id,
        "broker": broker,
        "account_name": account_name,
        "code": code,
        "name": name,
        "quantity": to_number(raw.get("quantity")),
        "avg_price": to_number(raw.get("avg_price")),
        "current_price": to_number(raw.get("current_price")),
        "currency": currency,
        "market": normalize_market(raw.get("market"), code, currency),
        "source": source,
        "sector": clean_text(raw.get("sector")) or get_default_sector(code, name),
        "price_updated_at": raw.get("price_updated_at"),
    }


def upsert_holdings(data: dict[str, Any], holdings: Iterable[dict[str, Any]], replace_source: str | None = None) -> int:
    if replace_source:
        data["holdings"] = [h for h in data["holdings"] if h.get("source") != replace_source]
    index = {(h["account_id"], h["code"]): position for position, h in enumerate(data["holdings"])}
    count = 0
    for holding in holdings:
        key = (holding["account_id"], holding["code"])
        if not holding["code"] or holding["quantity"] <= 0:
            continue
        if key in index:
            holding["id"] = data["holdings"][index[key]]["id"]
            data["holdings"][index[key]] = holding
        else:
            index[key] = len(data["holdings"])
            data["holdings"].append(holding)
        count += 1
    return count


def get_dashboard(data: dict[str, Any] | None = None, username: str | None = None) -> dict[str, Any]:
    data = data or read_portfolio(username)
    fx_rates = {key.upper(): to_number(value, 1.0) for key, value in data["settings"]["fx_rates"].items()}
    fx_rates.setdefault("KRW", 1.0)
    usd_rate = fx_rates.get("USD", 1.0)

    cash_balances = data["settings"].get("cash_balances", {})
    toss_cash = data["settings"].get("toss_cash", {})

    accounts: dict[str, dict[str, Any]] = {}
    for a in data["accounts"]:
        acc_id = a["id"]
        acc_cash = cash_balances.get(acc_id)
        if not acc_cash and a.get("broker") == "토스증권" and toss_cash:
            for seq_val in toss_cash.values():
                if isinstance(seq_val, dict):
                    acc_cash = seq_val
                    break
        acc_cash = acc_cash or {}
        cash_krw = to_number(acc_cash.get("KRW"))
        cash_usd = to_number(acc_cash.get("USD"))
        cash_total_krw = cash_krw + (cash_usd * usd_rate)

        accounts[acc_id] = {
            **a,
            "stock_value_krw": 0.0,
            "cash_krw": cash_krw,
            "cash_usd": cash_usd,
            "cash_total_krw": cash_total_krw,
            "market_value_krw": cash_total_krw,
            "cost_value_krw": 0.0,
            "profit_krw": 0.0,
            "holding_count": 0,
        }

    period_cache_file = _get_user_dir(username) / "period_rates.json"
    period_rates_data: dict[str, dict[str, float]] = data.get("settings", {}).get("period_rates", {})
    if not period_rates_data and period_cache_file.exists():
        try:
            period_rates_data = json.loads(period_cache_file.read_text(encoding="utf-8"))
            data.setdefault("settings", {})["period_rates"] = period_rates_data
        except Exception:
            pass
    elif period_rates_data and not period_cache_file.exists():
        try:
            period_cache_file.parent.mkdir(parents=True, exist_ok=True)
            period_cache_file.write_text(json.dumps(period_rates_data, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass

    enriched: list[dict[str, Any]] = []
    total_stock_value = total_stock_cost = 0.0
    for holding in data["holdings"]:
        item = deepcopy(holding)
        rate = fx_rates.get(item["currency"], 1.0 if item["currency"] == "KRW" else 0.0)
        item["fx_rate"] = rate
        item["market_value_krw"] = item["quantity"] * item["current_price"] * rate
        item["cost_value_krw"] = item["quantity"] * item["avg_price"] * rate
        item["profit_krw"] = item["market_value_krw"] - item["cost_value_krw"]
        item["return_rate"] = item["profit_krw"] / item["cost_value_krw"] * 100 if item["cost_value_krw"] else 0
        code_sym = str(item.get("code", "")).strip().upper()
        name_sym = str(item.get("name", "")).strip().upper()
        daily_map = data.get("settings", {}).get("daily_price_changes", {})
        p_info = period_rates_data.get(code_sym) or period_rates_data.get(name_sym) or {}
        day_rate = p_info.get("1D")
        if day_rate is None or to_number(day_rate) == 0.0:
            map_rate = daily_map.get(code_sym)
            if map_rate is None:
                map_rate = daily_map.get(name_sym)
            if map_rate is not None:
                day_rate = map_rate

        final_day_rate = to_number(day_rate, 0.0)
        item["day_change_rate"] = final_day_rate
        item["period_changes"] = {
            "1D": final_day_rate,
            "1W": to_number(p_info.get("1W"), final_day_rate),
            "1M": to_number(p_info.get("1M"), final_day_rate),
            "YTD": to_number(p_info.get("YTD"), final_day_rate),
            "1Y": to_number(p_info.get("1Y"), final_day_rate),
            "TOTAL": item["return_rate"],
        }
        enriched.append(item)
        total_stock_value += item["market_value_krw"]
        total_stock_cost += item["cost_value_krw"]
        account = accounts.get(item["account_id"])
        if account:
            account["stock_value_krw"] += item["market_value_krw"]
            account["market_value_krw"] += item["market_value_krw"]
            account["cost_value_krw"] += item["cost_value_krw"]
            account["profit_krw"] += item["profit_krw"]
            account["holding_count"] += 1

    enriched.sort(key=lambda h: h["market_value_krw"], reverse=True)

    total_cash_krw = sum(a["cash_krw"] for a in accounts.values())
    total_cash_usd = sum(a["cash_usd"] for a in accounts.values())
    total_cash_krw_combined = total_cash_krw + (total_cash_usd * usd_rate)

    total_value = total_stock_value + total_cash_krw_combined
    profit = total_stock_value - total_stock_cost

    account_list = sorted(accounts.values(), key=lambda a: a["market_value_krw"], reverse=True)
    for account in account_list:
        account["weight"] = account["market_value_krw"] / total_value * 100 if total_value else 0

    today_str = datetime.now(timezone.utc).astimezone().date().isoformat()
    previous_value = 0.0
    previous_date = None

    # 1. asset_records.json에서 오늘 이전의 가장 최근 기록 조회
    try:
        rec_file = _get_user_dir(username) / "asset_records.json"
        if rec_file.exists():
            rec_data = json.loads(rec_file.read_text(encoding="utf-8"))
            past_records = [
                r for r in rec_data.get("records", [])
                if isinstance(r, dict)
                and (r.get("owner") or "모두") == "모두"
                and r.get("date")
                and r.get("date") < today_str
                and to_number(r.get("total_value_krw")) > 0
            ]
            if past_records:
                past_records.sort(key=lambda x: x["date"])
                last_rec = past_records[-1]
                previous_value = to_number(last_rec["total_value_krw"])
                previous_date = last_rec["date"]
    except Exception:
        pass

    # 2. 보유종목의 실제 당일 등락 금액 합산 계산
    holding_day_gain = 0.0
    has_day_rates = False
    for h in enriched:
        val = h["market_value_krw"]
        r = h.get("day_change_rate") or 0.0
        if r != 0 and (100 + r) > 0:
            has_day_rates = True
            holding_day_gain += val * (r / (100 + r))

    # 3. 만약 이전 자산기록이 없거나 daily_snapshot이 오래되어 오차가 큰 경우 보유종목 등락 합산으로 자동 보정
    if previous_value <= 0:
        if has_day_rates and total_value > holding_day_gain:
            previous_value = total_value - holding_day_gain
            previous_date = "전일"
        else:
            daily_snapshot = data.get("settings", {}).get("daily_snapshot", {})
            if daily_snapshot.get("date") and daily_snapshot.get("date") < today_str:
                previous_value = to_number(daily_snapshot.get("value_krw"))
                previous_date = daily_snapshot.get("date")

    day_change = (total_value - previous_value) if previous_value > 0 else None

    currency_summary: dict[str, dict[str, float]] = {
        "KRW": {
            "market_value": total_cash_krw,
            "market_value_krw": total_cash_krw,
            "stock_value_krw": 0.0,
            "cost_value_krw": 0.0,
            "cash": total_cash_krw,
        },
        "USD": {
            "market_value": total_cash_usd,
            "market_value_krw": total_cash_usd * usd_rate,
            "stock_value": 0.0,
            "stock_value_krw": 0.0,
            "cost_value_krw": 0.0,
            "cash": total_cash_usd,
        },
    }

    classifications: dict[str, dict[str, Any]] = {}
    kr_etf_prefixes = ("KODEX", "TIGER", "ACE", "SOL", "PLUS", "RISE", "HANARO", "KOSEF", "ARIRANG", "KOACT", "WON", "1Q", "KIWOOM", "TIMEFOLIO", "WOORI", "KBSTAR")
    overseas_keywords = ("미국", "S&P", "나스닥", "NASDAQ", "다우", "DOW", "글로벌", "GLOBAL", "차이나", "중국", "CHINA", "인도", "INDIA", "일본", "JAPAN", "TOPIX", "NIKKEI", "유로", "EURO", "베트남", "VIETNAM", "FANG", "필라델피아", "빅테크", "BIG TECH", "월드", "WORLD", "선진국", "신흥국", "MSCI", "유럽", "대만", "해외")
    us_etf_tickers = {
        "QQQ", "QQQM", "SPY", "VOO", "IVV", "TLT", "TQQQ", "QLD", "SOXL", "SOXS",
        "SQQQ", "SCHD", "JEPI", "JEPQ", "DIA", "IWM", "VNQ", "GLD", "SLV", "SPYG",
        "QNDX", "SMH", "XLK", "XLE", "XLF", "XLV", "XLY", "XLP", "XLI", "XLU",
        "XLRE", "XLB", "IEF", "SHY", "BND", "AGG", "VT", "VTI", "VXUS", "ARKK",
        "BIL", "SHV", "VGK", "EEM", "VWO", "HYG", "LQD", "JNK", "TMF", "UPRO",
        "SPXU", "LABU", "LABD", "NUGT", "DUST", "FNGU", "BULZ"
    }

    for item in enriched:
        currency = item["currency"]
        bucket = currency_summary.setdefault(currency, {
            "market_value": 0.0, "market_value_krw": 0.0, "cost_value_krw": 0.0, "stock_value_krw": 0.0, "stock_value": 0.0, "cash": 0.0,
        })
        bucket["market_value"] += item["quantity"] * item["current_price"]
        bucket["market_value_krw"] += item["market_value_krw"]
        bucket["stock_value_krw"] = bucket.get("stock_value_krw", 0.0) + item["market_value_krw"]
        if currency == "USD":
            bucket["stock_value"] = bucket.get("stock_value", 0.0) + (item["quantity"] * item["current_price"])
        bucket["cost_value_krw"] += item["cost_value_krw"]

        name = item["name"].strip()
        name_upper = name.upper()
        code_upper = str(item.get("code", "")).strip().upper()

        if item["currency"] == "KRW":
            is_kr_etf = any(name_upper.startswith(p) for p in kr_etf_prefixes) or "ETF" in name_upper
            if is_kr_etf:
                if any(k.upper() in name_upper for k in overseas_keywords):
                    group = "국내상장해외ETF"
                else:
                    group = "국내ETF"
            else:
                group = "국내주식"
        else:
            if (
                code_upper in us_etf_tickers
                or "ETF" in name_upper
                or "TRUST" in name_upper
                or "FUND" in name_upper
                or "ISHARES" in name_upper
                or "VANGUARD" in name_upper
                or "INVESCO" in name_upper
                or "SPDR" in name_upper
            ):
                group = "해외ETF"
            else:
                group = "해외주식"

        classification = classifications.setdefault(group, {"name": group, "market_value_krw": 0.0, "cost_value_krw": 0.0, "holding_count": 0})
        classification["market_value_krw"] += item["market_value_krw"]
        classification["cost_value_krw"] += item["cost_value_krw"]
        classification["holding_count"] += 1

    if total_cash_krw_combined > 0:
        classifications["현금·예수금"] = {
            "name": "현금·예수금",
            "market_value_krw": total_cash_krw_combined,
            "cost_value_krw": total_cash_krw_combined,
            "profit_krw": 0.0,
            "return_rate": 0.0,
            "holding_count": 0,
            "weight": total_cash_krw_combined / total_value * 100 if total_value else 0.0,
        }

    classification_list = []
    for classification in classifications.values():
        classification.setdefault("profit_krw", classification["market_value_krw"] - classification["cost_value_krw"])
        classification.setdefault("return_rate", classification["profit_krw"] / classification["cost_value_krw"] * 100 if classification["cost_value_krw"] else 0.0)
        classification["weight"] = classification["market_value_krw"] / total_value * 100 if total_value else 0.0
        classification_list.append(classification)

    order_map = {
        "국내주식": 1,
        "국내ETF": 2,
        "국내상장해외ETF": 3,
        "해외주식": 4,
        "해외ETF": 5,
        "현금·예수금": 6,
    }
    classification_list.sort(key=lambda item: order_map.get(item["name"], 99))

    # 섹터별 포트폴리오 집계
    sectors: dict[str, dict[str, Any]] = {}
    for item in enriched:
        sec = item.get("sector") or get_default_sector(item.get("code", ""), item.get("name", ""))
        item["sector"] = sec
        s_obj = sectors.setdefault(sec, {
            "name": sec,
            "market_value_krw": 0.0,
            "cost_value_krw": 0.0,
            "profit_krw": 0.0,
            "return_rate": 0.0,
            "holding_count": 0,
            "weight": 0.0,
        })
        s_obj["market_value_krw"] += item["market_value_krw"]
        s_obj["cost_value_krw"] += item["cost_value_krw"]
        s_obj["profit_krw"] += item["profit_krw"]
        s_obj["holding_count"] += 1

    if total_cash_krw_combined > 0:
        sectors["현금·예수금"] = {
            "name": "현금·예수금",
            "market_value_krw": total_cash_krw_combined,
            "cost_value_krw": total_cash_krw_combined,
            "profit_krw": 0.0,
            "return_rate": 0.0,
            "holding_count": len([a for a in account_list if a.get("cash_total_krw", 0) > 0]),
            "weight": total_cash_krw_combined / total_value * 100 if total_value else 0.0,
        }

    sector_list = []
    for s_obj in sectors.values():
        s_obj["return_rate"] = s_obj["profit_krw"] / s_obj["cost_value_krw"] * 100 if s_obj["cost_value_krw"] else 0.0
        s_obj["weight"] = s_obj["market_value_krw"] / total_value * 100 if total_value else 0.0
        sector_list.append(s_obj)
    sector_list.sort(key=lambda x: x["market_value_krw"], reverse=True)

    return {
        "summary": {
            "total_value_krw": total_value,
            "total_stock_value_krw": total_stock_value,
            "total_cash_krw": total_cash_krw_combined,
            "cash_krw": total_cash_krw,
            "cash_usd": total_cash_usd,
            "total_cost_krw": total_stock_cost,
            "profit_krw": profit,
            "return_rate": profit / total_stock_cost * 100 if total_stock_cost else 0,
            "holding_count": len(enriched),
            "account_count": len(account_list),
        },
        "day_change": {
            "date": previous_date,
            "value_krw": previous_value,
            "change_krw": day_change,
            "change_rate": day_change / previous_value * 100 if day_change is not None and previous_value else None,
        },
        "accounts": account_list,
        "holdings": enriched,
        "fx_rates": fx_rates,
        "fx_info": data["settings"].get("fx_info", {}),
        "currency_summary": currency_summary,
        "classifications": classification_list,
        "sector_classifications": sector_list,
        "updated_at": data["updated_at"],
    }


def find_column(row: dict[str, Any], field: str) -> Any:
    normalized = {clean_text(key).lower(): value for key, value in row.items() if key is not None}
    for alias in FIELD_ALIASES[field]:
        if alias.lower() in normalized:
            return normalized[alias.lower()]
    return None


def rows_from_upload(filename: str, contents: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [clean_text(value) for value in values[0]]
        return [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return list(csv.DictReader(io.StringIO(contents.decode(encoding))))
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV 인코딩을 읽을 수 없습니다. UTF-8 또는 CP949 파일을 사용하세요.")


def import_rows(filename: str, contents: bytes, default_broker: str = "기타 증권사", username: str | None = None) -> tuple[int, list[str]]:
    rows = rows_from_upload(filename, contents)
    data = read_portfolio(username)
    normalized: list[dict[str, Any]] = []
    errors: list[str] = []
    for number, row in enumerate(rows, start=2):
        raw_code = clean_text(find_column(row, "code"))
        raw_name = clean_text(find_column(row, "name"))
        quantity = to_number(find_column(row, "quantity"))
        if (not raw_code and not raw_name) or quantity <= 0:
            errors.append(f"{number}행: 종목코드/종목명 또는 보유수량이 없어 건너뛰었습니다.")
            continue
        owner = clean_text(find_column(row, "owner")) or "모두"
        broker = clean_text(find_column(row, "broker")) or default_broker
        account_name = clean_text(find_column(row, "account")) or f"{broker} 가져온 계좌"
        account_id = get_or_add_account(data, broker, account_name, family_group=owner, source="import")
        
        # 계좌 소유자 필드 갱신
        for acc in data.get("accounts", []):
            if acc.get("id") == account_id:
                if owner and owner != "모두":
                    acc["owner"] = owner
                    acc["family_group"] = owner
                elif not acc.get("owner"):
                    acc["owner"] = "모두"
                    acc["family_group"] = "All"
                break

        h_item = normalize_holding({
            "code": raw_code, "name": raw_name, "quantity": quantity,
            "avg_price": find_column(row, "avg_price"), "current_price": find_column(row, "current_price"),
            "currency": find_column(row, "currency"), "market": find_column(row, "market"),
        }, account_id, broker, account_name, "import")
        h_item["owner"] = owner
        normalized.append(h_item)
    count = upsert_holdings(data, normalized)
    write_portfolio(data, username)
    return count, errors[:10]


def seed_demo() -> None:
    data = deepcopy(EMPTY_PORTFOLIO)
    data["settings"]["fx_rates"]["USD"] = 1350.0
    data["settings"]["fx_info"] = {"source": "예시 데이터", "valid_until": None}
    kb = get_or_add_account(data, "KB증권", "KB 국내 주식", "demo")
    toss = get_or_add_account(data, "토스증권", "토스 해외 주식", "demo")
    namoo = get_or_add_account(data, "NH투자증권(나무)", "나무 국내 주식", "demo")
    data["settings"]["cash_balances"] = {
        toss: {"KRW": 5000000.0, "USD": 2500.0},
        namoo: {"KRW": 1200000.0, "USD": 0.0},
    }
    examples = [
        {"code": "005930", "name": "삼성전자", "quantity": 12, "avg_price": 70200, "current_price": 72400, "currency": "KRW", "market": "KRX", "account_id": kb, "broker": "KB증권", "account_name": "KB 국내 주식"},
        {"code": "000660", "name": "SK하이닉스", "quantity": 4, "avg_price": 178000, "current_price": 198700, "currency": "KRW", "market": "KRX", "account_id": namoo, "broker": "NH투자증권(나무)", "account_name": "나무 국내 주식"},
        {"code": "AAPL", "name": "Apple", "quantity": 6, "avg_price": 195, "current_price": 214, "currency": "USD", "market": "NAS", "account_id": toss, "broker": "토스증권", "account_name": "토스 해외 주식"},
        {"code": "NVDA", "name": "NVIDIA", "quantity": 5, "avg_price": 118, "current_price": 128, "currency": "USD", "market": "NAS", "account_id": toss, "broker": "토스증권", "account_name": "토스 해외 주식"},
    ]
    upsert_holdings(data, [normalize_holding(item, item["account_id"], item["broker"], item["account_name"], "demo") for item in examples])
    write_portfolio(data)


def clear_portfolio() -> None:
    write_portfolio(deepcopy(EMPTY_PORTFOLIO))
